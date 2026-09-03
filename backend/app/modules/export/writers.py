"""Writing a bundle, and reading one back.

Formats are deliberately thin: the shaping is done by the time anything here
runs, so a writer is a rendering of `Table` and nothing decides anything. That
is what lets Stata and SPSS be added without any of the export's meaning moving
into them.

**CSV is written UTF-8 with a BOM.** Without one Excel reads a UTF-8 file as the
system code page, and a product that is RTL and Swahili from the start cannot
ship an export whose names are mojibake the moment a customer double-clicks it.
Every reader that matters strips it; `read_csv` below decodes `utf-8-sig`.

`read_csv` exists to make the round-trip invariant checkable — export, read
back, compare — rather than because anything imports CSVs. It is the writer's
inverse and it lives beside it for that reason: a reader in the test suite would
be a second reading of the format, and the two would drift.
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from typing import Literal

from .cells import Cell
from .manifest import Manifest
from .shape import Table

type Format = Literal["csv", "xlsx"]

#: Excel's own limit, not ours.
SHEET_NAME_LIMIT = 31


@dataclass(frozen=True)
class Bundle:
    """What an export is: named files, the tables in them, and the manifest."""

    files: tuple[tuple[str, bytes], ...]
    tables: tuple[Table, ...]
    manifest: Manifest

    def to_zip(self) -> bytes:
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, content in self.files:
                archive.writestr(name, content)
        return buffer.getvalue()


def write_bundle(
    tables: Sequence[Table], manifest: Manifest, *, fmt: Format
) -> Bundle:
    files: tuple[tuple[str, bytes], ...]
    if fmt == "xlsx":
        files = ((f"{manifest.form_id}.xlsx", write_xlsx(tables, manifest)),)
    else:
        files = tuple((f"{table.name}.csv", write_csv(table)) for table in tables) + (
            (
                "manifest.json",
                json.dumps(manifest.as_dict(), indent=2, ensure_ascii=False).encode(),
            ),
        )
    return Bundle(files=files, tables=tuple(tables), manifest=manifest)


def write_csv(table: Table) -> bytes:
    out = io.StringIO(newline="")
    writer = csv.writer(out, lineterminator="\r\n")
    writer.writerow([column.name for column in table.columns])
    for row in table.rows:
        writer.writerow(["" if cell is None else cell for cell in row])
    return out.getvalue().encode("utf-8-sig")


def read_csv(data: bytes) -> tuple[tuple[str, ...], tuple[tuple[str, ...], ...]]:
    """Header and rows, as text. Typing them back is `readback`'s job."""
    rows = list(csv.reader(io.StringIO(data.decode("utf-8-sig"), newline="")))
    if not rows:
        return (), ()
    return tuple(rows[0]), tuple(tuple(row) for row in rows[1:])


def write_xlsx(tables: Sequence[Table], manifest: Manifest) -> bytes:
    """One workbook, one sheet per table, plus the manifest as a sheet.

    The manifest sheet is not a duplicate of `manifest.json`: an .xlsx travels
    on its own — it is the file somebody emails — so an export whose separate
    manifest stayed behind would be exactly the silent, partly-unreadable file
    item 5 says must not exist.
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    book = Workbook()
    book.remove(book.active)

    used: set[str] = set()
    for table in tables:
        sheet = book.create_sheet(_sheet_name(table.name, used))
        sheet.append([column.name for column in table.columns])
        for row in table.rows:
            sheet.append(list(row))
        sheet.freeze_panes = "A2"
        for index in range(1, len(table.columns) + 1):
            sheet.column_dimensions[get_column_letter(index)].width = 18

    notes = book.create_sheet(_sheet_name("_manifest", used))
    notes.append(["Form", manifest.form_id, manifest.form_title or ""])
    notes.append(["Versions", ", ".join(str(v) for v in manifest.form_versions)])
    notes.append(["Shape", manifest.shape])
    notes.append(["Exported at", manifest.exported_at.isoformat()])
    notes.append(["Submissions", manifest.submission_count])
    notes.append([])
    for note in manifest.notes:
        notes.append(["Note", note])
    notes.append([])
    notes.append(
        [
            "table",
            "column",
            "label",
            "path",
            "dataType",
            "versions",
            "unreadable",
            "openableBy",
            "relevanceUncertain",
        ]
    )
    for table_manifest in manifest.tables:
        for column in table_manifest.columns:
            notes.append(
                [
                    table_manifest.name,
                    column.column,
                    column.label or "",
                    column.path or "",
                    column.data_type or "",
                    ", ".join(str(v) for v in column.versions),
                    column.unreadable or "",
                    ", ".join(column.openable_by),
                    "yes" if column.relevance_uncertain else "",
                ]
            )
    notes.column_dimensions["A"].width = 24
    notes.column_dimensions["B"].width = 28
    notes.column_dimensions["C"].width = 40

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def read_xlsx(data: bytes) -> Mapping[str, tuple[tuple[str, ...], tuple[tuple[Cell, ...], ...]]]:
    """Sheet name -> (header, rows). The manifest sheet is not returned."""
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    found: dict[str, tuple[tuple[str, ...], tuple[tuple[Cell, ...], ...]]] = {}
    for sheet in book.worksheets:
        if sheet.title.startswith("_manifest"):
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = tuple(str(cell) for cell in rows[0])
        found[sheet.title] = (
            header,
            tuple(tuple(row[: len(header)]) for row in rows[1:]),
        )
    return found


def _sheet_name(wanted: str, used: set[str]) -> str:
    """Excel's 31 characters, kept unique. Long form ids collide otherwise."""
    name = wanted[:SHEET_NAME_LIMIT]
    serial = 1
    while name in used:
        serial += 1
        suffix = f"~{serial}"
        name = wanted[: SHEET_NAME_LIMIT - len(suffix)] + suffix
    used.add(name)
    return name

