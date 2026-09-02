"""Reading an XLSForm workbook, and accounting for every cell in it.

## Why there is a ledger

An importer translates between two languages of different size, and every test
anyone writes for one compares "this input produced that output". Such a test
can only fail on an input somebody thought of. The construct nobody thought of
produces no output *and no test* — and the result still looks fine, because a
form with a question missing is a perfectly valid form. Nothing downstream can
tell: the IR compiles, both engines agree, the vectors pass, it publishes and
deploys and syncs. That is the silent drop, and no amount of test-writing
closes it, because the gap is exactly the set of things not thought of.

So this module does not test for it. It makes it impossible to express.

Every non-empty cell read from the workbook is registered here, and must end up
in exactly one of two sets:

  consumed  it produced IR
  reported  a diagnostic names it — sheet, row, column, value

At the end of an import, [CoverageLedger.residue] is whatever is in neither. If
it is not empty the import **fails**, loudly, as an internal error. A construct
that is neither translated nor reported cannot reach the author's report as
silence; it can only reach the developer as a crash.

That is the same move as `ServerConfig` and `FormCatalog.compiledFormForSubmission`
(breaks 30, 35): the way to stop something being got wrong is to stop it being
expressible, not to test that it was got right this time.

## What the ledger is blind to, and it is worth saying

It answers "was everything present accounted for". It cannot answer "was
anything present at all", and the two are not the same question: an empty
`survey` sheet registers no cells, so the residue is empty and the invariant is
perfectly satisfied by a workbook containing nothing.

That is not hypothetical. The official ODK XLSForm Template — a blank template,
499 rows by Excel's reckoning and not one with content — imported to a valid,
compilable form with zero questions, and every check in this repository passed
it. So emptiness is asked about directly instead, and a form with no questions
is refused by `forms.service.check_publishable` rather than only noted in a
report somebody may not read.

## Why cells and not rows

Row-level accounting would miss the commonest real loss, which is not a dropped
question but a dropped *column* on a question that imported. A row whose
`relevant` was silently ignored still produces a node, still compiles, and asks
a question that should have been hidden. The ledger is per cell so that
`survey!H14` can be reported even though `survey!14` was fine.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

# XLSForm's own sheet names. Everything else in the workbook is somebody's
# working notes — the ODK template ships eight documentation sheets with emoji
# in their names — and is reported as ignored rather than silently skipped, so
# an author who put questions on a sheet called "survey v2" finds out.
SURVEY_SHEET = "survey"
CHOICES_SHEET = "choices"
SETTINGS_SHEET = "settings"
KNOWN_SHEETS = (SURVEY_SHEET, CHOICES_SHEET, SETTINGS_SHEET)


@dataclass(frozen=True, order=True)
class CellRef:
    """Where something was, in the terms the author sees in Excel.

    `row` is 1-based and counts the header, so it is the number in the
    spreadsheet's own margin — not an index into a list. Somebody reading the
    report has the file open; the reference has to match what is in front of
    them.
    """

    sheet: str
    row: int
    column: str

    def __str__(self) -> str:
        # Written the way somebody with the file open would say it, because
        # that is who reads the report. "survey!type6" looks like a cell
        # reference and is not one.
        if self.column:
            return f"{self.sheet} row {self.row}, column '{self.column}'"
        return f"{self.sheet} row {self.row}"


@dataclass
class Cell:
    ref: CellRef
    value: str
    #: Excel's column letter, for a reader matching this against the file.
    letter: str


@dataclass
class Row:
    """One spreadsheet row, addressed by column *name* rather than position."""

    sheet: str
    number: int
    cells: dict[str, Cell]

    def get(self, column: str) -> str | None:
        cell = self.cells.get(column)
        return cell.value if cell else None

    def ref(self, column: str) -> CellRef:
        cell = self.cells.get(column)
        return cell.ref if cell else CellRef(self.sheet, self.number, column)

    @property
    def is_empty(self) -> bool:
        return not self.cells


class CoverageLedger:
    """Every non-empty cell, and what became of it.

    The invariant is checked by [residue], not by trust: a cell registered and
    never resolved is a hole, and the import refuses rather than reporting a
    form that quietly lost something.
    """

    def __init__(self) -> None:
        self._known: set[CellRef] = set()
        self._consumed: set[CellRef] = set()
        self._reported: set[CellRef] = set()

    def register(self, ref: CellRef) -> None:
        """Record that the workbook had something here."""
        self._known.add(ref)

    def consume(self, ref: CellRef) -> None:
        """This cell produced IR."""
        self._consumed.add(ref)

    def report(self, ref: CellRef) -> None:
        """A diagnostic names this cell."""
        self._reported.add(ref)

    @property
    def residue(self) -> list[CellRef]:
        """Cells that produced nothing and were never mentioned.

        Sorted, because the failure message is read by a developer who needs to
        find the first one, and set iteration order would make the same bug
        look different on every run.
        """
        return sorted(self._known - self._consumed - self._reported)

    @property
    def counts(self) -> dict[str, int]:
        return {
            "cells": len(self._known),
            "consumed": len(self._consumed),
            "reported": len(self._reported),
        }


class WorkbookError(Exception):
    """The file is not a readable XLSForm workbook at all.

    Distinct from a diagnostic: a diagnostic is something *about* a form, and
    this is the absence of one. A .docx renamed to .xlsx has no sheet to point a
    row number at.
    """


def _column_letter(index: int) -> str:
    """0-based column index to Excel's A, B, ... Z, AA."""
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _clean(value: Any) -> str | None:
    """A cell's text, or None if it holds nothing worth reading.

    Excel hands back floats for integers a person typed as integers — a
    `version` of 2 arrives as 2.0, and a form version of "2.0" is not a form
    version. Trailing `.0` is dropped for exactly that reason and no other.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value)
    text = text.strip()
    return text or None


@dataclass
class Sheet:
    name: str
    columns: list[str]
    rows: list[Row]
    #: Header cells whose text was blank but which sat between named columns.
    unnamed_columns: list[str] = field(default_factory=list)


@dataclass
class Workbook:
    sheets: dict[str, Sheet]
    #: Sheets XLSForm does not define. Reported, never silently skipped.
    extra_sheets: list[str]
    ledger: CoverageLedger

    def sheet(self, name: str) -> Sheet | None:
        return self.sheets.get(name)


def read(data: bytes, ledger: CoverageLedger) -> Workbook:
    """Read an .xlsx into rows addressed by column name, registering every cell.

    Only the three XLSForm sheets are read into rows. The rest are recorded by
    name so the report can say they were ignored — the ODK template carries
    eight documentation sheets, and an author needs to know we skipped
    "👋 Start here" on purpose but would also need to know if we skipped a
    sheet holding their questions.
    """
    try:
        import openpyxl
    except ModuleNotFoundError as exc:  # pragma: no cover - dependency is declared
        raise WorkbookError("openpyxl is not installed; cannot read .xlsx") from exc

    import io
    import warnings

    try:
        with warnings.catch_warnings():
            # openpyxl warns about Excel features it drops (data validation,
            # conditional formatting). None of them carry form content, and the
            # warning would otherwise land in an API response.
            warnings.simplefilter("ignore")
            book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except WorkbookError:
        raise
    except Exception as exc:
        raise WorkbookError(f"this file could not be opened as a .xlsx workbook: {exc}") from exc

    sheets: dict[str, Sheet] = {}
    extra: list[str] = []
    try:
        for worksheet in book.worksheets:
            name = str(worksheet.title).strip()
            # Case-insensitively, because real workbooks say "Survey".
            canonical = next((k for k in KNOWN_SHEETS if k == name.lower()), None)
            if canonical is None:
                extra.append(name)
                continue
            sheets[canonical] = _read_sheet(worksheet, canonical, ledger)
    finally:
        book.close()

    if SURVEY_SHEET not in sheets:
        raise WorkbookError(
            "this workbook has no 'survey' sheet, so it is not an XLSForm. "
            f"Sheets found: {', '.join(extra) or 'none'}"
        )

    return Workbook(sheets=sheets, extra_sheets=extra, ledger=ledger)


#: Rows and columns beyond which a sheet is not a form any more.
#:
#: Both are defences against the declared dimension rather than the data. A
#: workbook's XML states its own size and openpyxl's read-only mode believes it,
#: so a sheet claiming three million rows yields three million tuples of empty
#: cells — the UCL biomass form, 81 real rows, took 37 seconds and made a
#: billion calls before this existed. On an endpoint that accepts uploads that
#: is not only slow, it is a way to keep a worker busy with a small file.
_MAX_ROWS = 20_000
_MAX_COLUMNS = 500


def _read_sheet(worksheet: Any, name: str, ledger: CoverageLedger) -> Sheet:
    # Ignore the dimension the file declares and let openpyxl work it out from
    # the cells that are actually there. Without this the loop below is sized
    # by a number the uploader chose.
    if hasattr(worksheet, "reset_dimensions"):
        worksheet.reset_dimensions()

    raw: list[tuple[Any, ...]] = []
    blank_run = 0
    for row in worksheet.iter_rows(values_only=True):
        raw.append(row[:_MAX_COLUMNS])
        # A long tail of empty rows is what a spreadsheet looks like after
        # somebody deletes content; the sheet keeps the size it once had.
        blank_run = 0 if any(c is not None for c in row[:_MAX_COLUMNS]) else blank_run + 1
        if blank_run >= 200 or len(raw) >= _MAX_ROWS:
            break
    while raw and not any(c is not None for c in raw[-1]):
        raw.pop()
    if not raw:
        return Sheet(name=name, columns=[], rows=[])

    header = raw[0]
    columns: list[str] = []
    letters: list[str] = []
    unnamed: list[str] = []
    for index, cell in enumerate(header):
        text = _clean(cell)
        letter = _column_letter(index)
        if text is None:
            columns.append("")
            letters.append(letter)
            continue
        # XLSForm column names are case-insensitive and carry stray spaces in
        # real files — the ODK template ships a column literally named "audio ".
        normalised = re.sub(r"\s+", " ", text).strip().lower()
        columns.append(normalised)
        letters.append(letter)

    rows: list[Row] = []
    for offset, values in enumerate(raw[1:], start=2):
        cells: dict[str, Cell] = {}
        for index, value in enumerate(values):
            text = _clean(value)
            if text is None:
                continue
            column = columns[index] if index < len(columns) else ""
            letter = letters[index] if index < len(letters) else _column_letter(index)
            if not column:
                # A value under a blank header. Registered so it cannot be lost
                # silently; the importer reports it as an unreadable column.
                column = f"<unnamed column {letter}>"
                if column not in unnamed:
                    unnamed.append(column)
            ref = CellRef(sheet=name, row=offset, column=column)
            ledger.register(ref)
            cells[column] = Cell(ref=ref, value=text, letter=letter)
        rows.append(Row(sheet=name, number=offset, cells=cells))

    return Sheet(name=name, columns=[c for c in columns if c], rows=rows, unnamed_columns=unnamed)
