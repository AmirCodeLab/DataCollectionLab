"""XLSForm import: what survives, what is reported, and what cannot be silent.

The most important test in here is `test_a_dropped_cell_is_an_error_not_a_silence`.
Everything else checks that a particular construct is translated the way it
should be, and every one of those can only fail on a construct somebody thought
of. That one checks the property that holds for constructs nobody thought of.

Workbooks are built in memory rather than committed as fixtures. A fixture is a
form written to suit the importer, and the whole lesson of this work is that
those prove less than they appear to — the real forms are exercised separately
by `test_xlsform_corpus.py`.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl
import pytest

from app.modules.forms.xlsform import workbook as workbook_module
from app.modules.forms.xlsform.importer import CoverageHole, ImportFailed, import_workbook


def build(
    survey: list[list[Any]],
    choices: list[list[Any]] | None = None,
    settings: list[list[Any]] | None = None,
) -> bytes:
    """An .xlsx from rows, first row of each block being its header."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "survey"
    for row in survey:
        sheet.append(row)
    if choices:
        page = book.create_sheet("choices")
        for row in choices:
            page.append(row)
    if settings:
        page = book.create_sheet("settings")
        for row in settings:
            page.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


SIMPLE = [
    ["type", "name", "label"],
    ["text", "who", "Who are you?"],
    ["integer", "age", "How old?"],
]


def diagnostics_by_code(result, code: str):
    return [d for d in result.diagnostics if d.code == code]


# -- the property that survives constructs nobody thought of -----------------


def test_a_dropped_cell_is_an_error_not_a_silence(monkeypatch: pytest.MonkeyPatch) -> None:
    """A cell that produces nothing and is never reported fails the import.

    This is the guard the whole module is built around, so it is tested by
    breaking it: `_consume_remaining` is what accounts for every column nothing
    else claimed, and with it disabled an unrecognised column is exactly the
    silent drop this design exists to prevent — the form still imports, still
    compiles, still looks complete, and a column is gone.

    Without the ledger that would be caught by nothing. With it, it cannot even
    return.
    """
    data = build(
        [
            ["type", "name", "label", "some_future_column"],
            ["text", "who", "Who are you?", "a value nothing here understands"],
        ]
    )

    # Sanity first: normally that column is reported, and the import succeeds.
    result = import_workbook(data)
    assert diagnostics_by_code(result, "unknown_column"), (
        "an unrecognised column should be reported"
    )

    # Now break the accounting and confirm the import refuses rather than
    # returning a form that quietly lost the cell.
    from app.modules.forms.xlsform import importer as importer_module

    monkeypatch.setattr(
        importer_module._Importer, "_consume_remaining", lambda *a, **k: None
    )
    with pytest.raises(CoverageHole) as hole:
        import_workbook(data)
    assert "some_future_column" in str(hole.value)
    assert "importer bug" in str(hole.value)


def test_the_ledger_cannot_see_an_empty_workbook() -> None:
    """The blind spot, asserted so it stays known rather than becoming a surprise.

    Coverage answers "was everything present accounted for". It cannot answer
    "was anything present at all": an empty survey sheet registers no cells, so
    the residue is empty and the invariant is satisfied by nothing. The official
    ODK XLSForm Template is exactly this, and it imported to a valid compilable
    form that passed every check in the repository.

    So emptiness is checked separately — here, and at publish
    (`forms.service.check_publishable`), which is the gate that matters.
    """
    result = import_workbook(build([["type", "name", "label"]]))

    assert result.coverage["cells"] == 0, "an empty sheet has nothing to account for"
    assert not result.coverage["reported"], "and so nothing is reported by coverage"
    # The ledger is satisfied. The direct question is what catches it.
    assert result.questions == 0
    assert not result.publishable
    assert diagnostics_by_code(result, "no_questions")


# -- refusals ---------------------------------------------------------------


def test_a_file_that_is_not_a_workbook_is_refused() -> None:
    with pytest.raises(ImportFailed):
        import_workbook(b"this is not a spreadsheet")


def test_a_workbook_with_no_survey_sheet_is_refused() -> None:
    book = openpyxl.Workbook()
    book.active.title = "data"
    buffer = io.BytesIO()
    book.save(buffer)
    with pytest.raises(ImportFailed, match="no 'survey' sheet"):
        import_workbook(buffer.getvalue())


# -- expressions: the severity rule -----------------------------------------


def test_an_untranslatable_relevant_is_an_error_not_a_warning() -> None:
    """Dropped, it would show a conditional question to everybody.

    That is the difference between the two severities: a warning is something
    the author may ignore, and ignoring this changes who is asked what.
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "relevant"],
                ["text", "a", "A", None],
                ["text", "b", "B", "atan(${a}) > 1"],
            ]
        )
    )
    failures = diagnostics_by_code(result, "untranslatable_expression")
    assert len(failures) == 1
    assert failures[0].severity == "error"
    assert "shown to everybody" in failures[0].message
    assert not result.publishable
    assert result.instrumentation.unsupported_functions == {"atan": 1}


def test_an_untranslatable_default_is_only_a_warning() -> None:
    # The question is still asked and still validated; it starts empty.
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "default"],
                ["text", "a", "A", "atan(${b})"],
                ["text", "b", "B", None],
            ]
        )
    )
    failures = diagnostics_by_code(result, "untranslatable_expression")
    assert [d.severity for d in failures] == ["warning"]
    assert result.publishable


def test_a_constraint_reads_dot_as_this_question() -> None:
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "constraint"],
                ["integer", "age", "Age", ". < 120"],
            ]
        )
    )
    node = result.form["children"][0]
    assert node["constraint"] == {
        "op": "lt",
        "args": [{"op": "ref", "path": "age"}, {"op": "lit", "value": 120}],
    }


def test_curly_quotes_are_read_as_quotes() -> None:
    """Excel autocorrects `'` to U+2019 without asking.

    Real forms are full of them — seven expressions in the UCL biomass survey —
    and refusing a file for something Excel did to it is refusing it for no
    reason the author can see.
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "relevant"],
                ["text", "a", "A", None],
                ["text", "b", "B", "${a} = ‘yes’"],
            ]
        )
    )
    assert not diagnostics_by_code(result, "untranslatable_expression")
    assert result.form["children"][1]["relevant"]["args"][1] == {"op": "lit", "value": "yes"}


# -- types ------------------------------------------------------------------


def test_a_type_in_the_spec_but_not_collectable_is_an_error() -> None:
    """It would deploy and arrive as a question nobody can answer.

    `barcode` is in Form IR §2.1 and both engines evaluate it; no client has a
    widget. That combination is the dangerous one, because everything except the
    phone says it is fine.
    """
    result = import_workbook(
        build([["type", "name", "label"], ["barcode", "code", "Scan it"]])
    )
    refusals = diagnostics_by_code(result, "type_not_collectable")
    assert len(refusals) == 1
    assert refusals[0].severity == "error"
    assert result.instrumentation.uncollectable_types == {"barcode": 1}


def test_an_unknown_type_names_the_question_it_lost() -> None:
    """Nothing recognised it, so "check the spelling" is honest here."""
    result = import_workbook(
        build([["type", "name", "label"], ["holograph", "h", "Wave"], ["text", "a", "A"]])
    )
    lost = diagnostics_by_code(result, "unknown_type")
    assert lost[0].severity == "error"
    assert lost[0].blame == "author"
    assert "`h` was not imported" in lost[0].message
    assert "spelling" in (lost[0].remedy or "")
    assert result.questions == 1


def test_a_real_xlsform_type_we_have_not_built_is_ours_not_the_author_s() -> None:
    """`select_one_from_file` is spelled correctly and is a real type.

    Telling somebody to check the spelling of a correctly spelled word is the
    same defect as a connect-timeout message asserting something is at the
    address: a conclusion the evidence does not support. Worse here, because
    acting on it costs an evening they cannot win.
    """
    result = import_workbook(
        build([["type", "name", "label"], ["select_one_from_file places.csv", "p", "Place"]])
    )
    lost = diagnostics_by_code(result, "type_not_implemented")
    assert lost[0].severity == "error"
    assert lost[0].blame == "platform"
    assert "valid XLSForm type" in lost[0].message
    assert "Nothing is wrong with your spreadsheet" in (lost[0].remedy or "")
    assert not diagnostics_by_code(result, "unknown_type")


def test_a_type_the_ir_cannot_express_says_it_needs_rewriting() -> None:
    # `trigger` has no IR equivalent and none is planned. Waiting for us is the
    # wrong advice; rewriting is the only advice.
    result = import_workbook(
        build([["type", "name", "label"], ["trigger", "t", "Tap to acknowledge"]])
    )
    lost = diagnostics_by_code(result, "unsupported_type")
    assert lost[0].blame == "author"
    assert "no equivalent planned" in (lost[0].remedy or "")


def test_a_reference_to_a_dropped_question_is_reported_under_its_cause() -> None:
    """One missing feature is one problem, not four.

    The UCL form had five of these — a `relevant` pointing at a
    `select_one_from_file` question a row or two above. Counting them
    separately reported one gap as six errors and told an author their form
    was a disaster.
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "relevant"],
                ["select_one_from_file places.csv", "place", "Place", None],
                ["text", "other", "Other", "${place} = 'x'"],
            ]
        )
    )
    cascade = diagnostics_by_code(result, "unknown_reference")
    assert len(cascade) == 1
    assert cascade[0].caused_by is not None, "should point at the drop that caused it"
    assert cascade[0].blame == "platform", "our missing feature, not the author's error"
    assert "see above" in cascade[0].message

    roots = [d for d in result.diagnostics if d.severity == "error" and d.caused_by is None]
    assert len(roots) == 1, f"one root cause, got {[d.code for d in roots]}"


def test_a_relevant_on_a_repeat_is_imported_not_reported_as_unknown() -> None:
    """Groups and repeats carry `relevant` too (§2.2, §2.3).

    This ran only for questions, so a `relevant` on a `begin repeat` fell
    through to the unknown-column branch — reported as "`relevant` is not a
    column this importer understands", which is false, and as a warning, which
    understates it. The engine inherits ancestor relevance, so the whole
    subtree lost its condition.
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "relevant"],
                ["select_one yn", "any", "Any?", None],
                ["begin repeat", "items", "Items", "${any} = 'y'"],
                ["text", "n", "Name", None],
                ["end repeat", None, None, None],
            ],
            choices=[["list_name", "name", "label"], ["yn", "y", "Yes"]],
        )
    )
    assert not diagnostics_by_code(result, "unknown_column"), (
        [d.message for d in diagnostics_by_code(result, "unknown_column")]
    )
    repeat = result.form["children"][1]
    assert repeat["type"] == "repeat"
    assert repeat["relevant"] == {
        "op": "eq",
        "args": [{"op": "ref", "path": "any"}, {"op": "lit", "value": "y"}],
    }


def test_the_legacy_select_spellings_are_read() -> None:
    # The ODK "widgets" sample still uses these throughout.
    result = import_workbook(
        build(
            [
                ["type", "name", "label"],
                ["select all that apply from yn", "s", "Pick"],
                ["select one from yn", "t", "Pick one"],
            ],
            choices=[["list_name", "name", "label"], ["yn", "y", "Yes"], ["yn", "n", "No"]],
        )
    )
    assert [c["dataType"] for c in result.form["children"]] == [
        "select_multiple",
        "select_one",
    ]
    assert len(diagnostics_by_code(result, "legacy_type_spelling")) == 2


def test_a_select_referring_to_a_missing_list_is_an_error() -> None:
    result = import_workbook(
        build([["type", "name", "label"], ["select_one colours", "c", "Colour"]])
    )
    missing = diagnostics_by_code(result, "missing_choice_list")
    assert missing[0].severity == "error"
    assert "colours" in missing[0].message


# -- identifiers ------------------------------------------------------------


def test_a_camel_case_name_is_renamed_and_the_rename_is_reported() -> None:
    """The id is what an export column is called, so a rename has to be told."""
    result = import_workbook(
        build([["type", "name", "label"], ["text", "numberAsString", "N"]])
    )
    renamed = diagnostics_by_code(result, "name_normalised")
    assert renamed[0].severity == "warning"
    assert result.form["children"][0]["id"] == "numberasstring"


def test_a_rename_is_followed_through_every_reference() -> None:
    """Otherwise `relevant` points at a name nothing answers.

    That would be caught — §4.2 makes an unresolvable reference a compile error
    — but caught is not the same as correct, and "we renamed it and followed it
    through" is a far better report than "your form no longer compiles".
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "relevant"],
                ["integer", "myAge", "Age", None],
                ["text", "b", "B", "${myAge} > 18"],
            ]
        )
    )
    assert result.form["children"][0]["id"] == "myage"
    assert result.form["children"][1]["relevant"]["args"][0] == {"op": "ref", "path": "myage"}
    assert not diagnostics_by_code(result, "unknown_reference")


def test_two_names_that_normalise_to_one_are_refused_not_invented_around() -> None:
    """Inventing a suffix would scramble which answers belong to which question."""
    result = import_workbook(
        build(
            [
                ["type", "name", "label"],
                ["text", "myField", "One"],
                ["text", "MyField", "Two"],
            ]
        )
    )
    collision = diagnostics_by_code(result, "name_collision_after_normalising")
    assert collision[0].severity == "error"
    assert result.questions == 1


def test_a_reference_to_nothing_is_reported_against_its_cell() -> None:
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "relevant"],
                ["text", "a", "A", "${nonexistent} = 'x'"],
            ]
        )
    )
    unknown = diagnostics_by_code(result, "unknown_reference")
    assert unknown[0].severity == "error"
    assert unknown[0].ref is not None
    assert unknown[0].ref.column == "relevant"
    assert "nonexistent" in unknown[0].message
    # ...and the expression is not left pointing at nothing.
    assert "relevant" not in result.form["children"][0]


# -- provenance -------------------------------------------------------------


def test_every_diagnostic_can_be_located_in_the_spreadsheet() -> None:
    """Sheet, row and column as fields, so a console can link to the cell.

    Some diagnostics are about the whole workbook and have no cell; those are
    the only ones allowed no location.
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label", "mystery"],
                ["text", "a", "A", "x"],
                ["holograph", "b", "B", None],
            ]
        )
    )
    located = [d for d in result.diagnostics if d.ref is not None]
    assert located, "nothing was located at all"
    for diagnostic in located:
        assert diagnostic.ref.sheet in ("survey", "choices", "settings")
        assert diagnostic.ref.row >= 2, "row 1 is the header"
        assert diagnostic.ref.column


def test_the_ir_it_returns_always_compiles_or_says_why() -> None:
    """The importer never hands back IR the engine would refuse in silence.

    Found by importing the UCL biomass survey, which nests a repeat inside a
    repeat: the importer produced IR that could not be compiled and called it
    publishable.
    """
    result = import_workbook(
        build(
            [
                ["type", "name", "label"],
                ["begin repeat", "outer", "Outer"],
                ["begin repeat", "inner", "Inner"],
                ["text", "a", "A"],
                ["end repeat", None, None],
                ["end repeat", None, None],
            ]
        )
    )
    refusal = diagnostics_by_code(result, "does_not_compile")
    assert refusal, "nested repeats should be reported, not returned as valid IR"
    assert refusal[0].severity == "error"
    assert not result.publishable


def test_an_unbalanced_group_is_reported() -> None:
    result = import_workbook(
        build([["type", "name", "label"], ["begin group", "g", "G"], ["text", "a", "A"]])
    )
    assert diagnostics_by_code(result, "unbalanced_group")
    assert not result.publishable


def test_sheets_that_are_not_xlsform_are_named_not_skipped_quietly() -> None:
    """The ODK template ships eight documentation sheets.

    Saying which were skipped is how an author finds out their questions were
    on a sheet called "survey v2".
    """
    book = openpyxl.Workbook()
    book.active.title = "survey"
    for row in SIMPLE:
        book.active.append(row)
    book.create_sheet("my working notes")
    buffer = io.BytesIO()
    book.save(buffer)

    result = import_workbook(buffer.getvalue())
    ignored = diagnostics_by_code(result, "sheet_ignored")
    assert any("my working notes" in d.message for d in ignored)


def test_a_value_under_a_blank_header_is_still_accounted_for() -> None:
    # The cell exists and means nothing; losing it silently is the failure.
    result = import_workbook(
        build([["type", "name", "label", None], ["text", "a", "A", "orphaned"]])
    )
    assert any("unnamed column" in (d.ref.column if d.ref else "") for d in result.diagnostics)


def test_cell_references_read_the_way_a_person_would_say_them() -> None:
    ref = workbook_module.CellRef(sheet="survey", row=14, column="relevant")
    assert str(ref) == "survey row 14, column 'relevant'"
